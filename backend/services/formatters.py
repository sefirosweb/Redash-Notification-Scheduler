from typing import List, Dict
from io import BytesIO
from jinja2 import Template
import openpyxl
from openpyxl.styles import Font


def rows_to_html(rows: List[Dict], columns: List[Dict], total_row: Dict = None) -> str:
    table_template = Template('''
    <table style="border-collapse:collapse;width:100%;font-family:sans-serif;">
        <thead>
            <tr>
                {% for col in columns %}
                    <th style="background:#6c3483;color:#fff;padding:8px;border:1px solid #ccc;">{{ col['name'] }}</th>
                {% endfor %}
            </tr>
        </thead>
        <tbody>
            {% for row in rows %}
            <tr>
                {% for col in columns %}
                    <td style="padding:8px;border:1px solid #ccc;">{{ row[col['name']] }}</td>
                {% endfor %}
            </tr>
            {% endfor %}
            {% if total_row %}
            <tr style="font-weight:bold;background:#eee;">
                {% for col in columns %}
                    <td style="padding:8px;border:1px solid #ccc;">{{ total_row[col['name']] }}</td>
                {% endfor %}
            </tr>
            {% endif %}
        </tbody>
    </table>
    ''')
    return table_template.render(rows=rows, columns=columns, total_row=total_row)


def rows_to_pdf(rows: List[Dict], columns: List[Dict], total_row: Dict = None) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib import colors

    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), leftMargin=20, rightMargin=20, topMargin=20, bottomMargin=20)

    headers = [col['name'] for col in columns]
    data = [headers]
    for row in rows:
        data.append([str(row.get(col['name'], '')) for col in columns])
    if total_row:
        data.append([str(total_row.get(col['name'], '')) for col in columns])

    col_count = len(headers)
    col_width = (landscape(A4)[0] - 40) / col_count if col_count else 100

    table = Table(data, colWidths=[col_width] * col_count, repeatRows=1)
    style = [
        ('BACKGROUND',  (0, 0), (-1, 0),  colors.HexColor('#6c3483')),
        ('TEXTCOLOR',   (0, 0), (-1, 0),  colors.white),
        ('FONTNAME',    (0, 0), (-1, 0),  'Helvetica-Bold'),
        ('FONTSIZE',    (0, 0), (-1, -1), 7),
        ('GRID',        (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('VALIGN',      (0, 0), (-1, -1), 'MIDDLE'),
        ('PADDING',     (0, 0), (-1, -1), 4),
    ]
    if total_row:
        style.append(('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'))
        style.append(('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#eeeeee')))
    table.setStyle(TableStyle(style))

    doc.build([table])
    return output.getvalue()


def rows_to_excel(rows: List[Dict], columns: List[Dict], total_row: Dict = None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    for idx, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=idx, value=col['name'])
        cell.font = Font(bold=True, color="6C3483")
    for r_idx, row in enumerate(rows, 2):
        for c_idx, col in enumerate(columns, 1):
            ws.cell(row=r_idx, column=c_idx, value=row.get(col['name'], ""))
    if total_row:
        t_idx = len(rows) + 2
        for c_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=t_idx, column=c_idx, value=total_row.get(col['name'], ""))
            cell.font = Font(bold=True)
    output = BytesIO()
    wb.save(output)
    return output.getvalue()
